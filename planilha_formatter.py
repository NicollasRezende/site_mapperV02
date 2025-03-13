import pandas as pd
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import logging
import argparse
import re
from urllib.parse import urlparse

# Configuração de logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("formatter.log"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger("PlanilhaFormatter")

class PlanilhaFormatter:
    """
    Classe para formatar e organizar os dados do mapeamento de sites,
    seguindo regras específicas de hierarquia e formatação.
    """
    
    def __init__(self, input_csv, output_dir="output", site_prefix=None):
        """
        Inicializa o formatador.
        
        Args:
            input_csv: Caminho para o CSV gerado pelo mapeador
            output_dir: Diretório para salvar o resultado
            site_prefix: Prefixo do site a substituir por "Raiz" (ex: "Tribunal Administrativo de Recursos Fiscais")
        """
        self.input_csv = input_csv
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)
        self.site_prefix = site_prefix  # Prefixo do site a substituir por "Raiz"
        
        # Cores para Excel
        self.BLUE_COLOR = "9FC5E8"
        self.YELLOW_COLOR = "FFE599"
        self.RED_COLOR = "EA9999"
        
        # Cores para linhas de cabeçalho
        self.BLUE_HEADER_COLOR = "CFE2F3"
        self.YELLOW_HEADER_COLOR = "FFFAEC"
        self.RED_HEADER_COLOR = "F4CCCC"
        
        # Cores para dados
        self.BLUE_LIGHT_COLOR = "EDF5FC"
        self.YELLOW_LIGHT_COLOR = "FFFAEC"
        self.RED_LIGHT_COLOR = "FFF2F2"
        
        # Data para nome do arquivo
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
    def process(self):
        """
        Processa o CSV, reorganiza e formata os dados.
        """
        logger.info(f"Processando arquivo CSV: {self.input_csv}")
        
        # Carregar dados
        try:
            df = pd.read_csv(self.input_csv)
            logger.info(f"Carregadas {len(df)} linhas do CSV")
        except Exception as e:
            logger.error(f"Erro ao carregar CSV: {str(e)}")
            return False
        
        # Detectar nome do site se não foi fornecido
        if not self.site_prefix:
            self._detect_site_name(df)
            
        logger.info(f"Nome do site a substituir: {self.site_prefix}")
            
        # Detectar e processar a Home Page
        df = self._process_home_page(df)
        
        # Processar breadcrumbs (substituindo o nome do site por "Raiz")
        # Importante: processar primeiro os breadcrumbs, depois as hierarquias
        df = self._process_breadcrumbs(df)
            
        # Processar hierarquias
        df = self._process_hierarchies(df)
        
        # Processar tipos de páginas
        df = self._process_page_types(df)
        
        # Reordenar linhas
        df = self._reorder_rows(df)
        
        # Salvar CSV processado
        processed_csv = f"{self.output_dir}/formatted_{self.timestamp}.csv"
        df.to_csv(processed_csv, index=False)
        logger.info(f"CSV processado salvo em: {processed_csv}")
        
        # Salvar Excel formatado
        excel_file = f"{self.output_dir}/formatted_{self.timestamp}.xlsx"
        self._save_to_excel(df, excel_file)
        logger.info(f"Excel formatado salvo em: {excel_file}")
        
        return True
    
    def _detect_site_name(self, df):
        """
        Detecta o nome do site a partir dos dados.
        """
        # Tentar detectar a partir do breadcrumb
        if 'Breadcrumb' in df.columns:
            for breadcrumb in df['Breadcrumb'].dropna():
                parts = breadcrumb.split(' > ')
                if parts:
                    self.site_prefix = parts[0]
                    logger.info(f"Nome do site detectado a partir do breadcrumb: {self.site_prefix}")
                    return
        
        # Tentar detectar a partir da hierarquia
        if 'Hierarquia' in df.columns:
            for hierarchy in df['Hierarquia'].dropna():
                parts = hierarchy.split(' > ')
                if len(parts) > 1:
                    self.site_prefix = parts[1]  # Assumindo que o formato é "Raiz > Nome do Site > ..."
                    logger.info(f"Nome do site detectado a partir da hierarquia: {self.site_prefix}")
                    return
                    
        # Se não conseguiu detectar, usar um valor padrão
        self.site_prefix = "Tribunal Administrativo de Recursos Fiscais"
        logger.warning(f"Não foi possível detectar o nome do site. Usando padrão: {self.site_prefix}")
    
    def _process_home_page(self, df):
        """
        Identifica e processa a página inicial (home).
        """
        logger.info("Identificando a página inicial...")
        
        # Detectar possíveis home pages por padrões de URL
        base_domain = None
        home_candidates = []
        
        for idx, url in enumerate(df['De']):
            if pd.isna(url):
                continue
                
            parsed = urlparse(url)
            # Encontrar o domínio base se ainda não temos
            if base_domain is None:
                base_domain = f"{parsed.scheme}://{parsed.netloc}"
                
            # Verificar se é a página raiz (/) ou index.html
            if (url == base_domain or 
                url == f"{base_domain}/" or 
                url.endswith('/index.html') or
                parsed.path == "/" or
                parsed.path == ""):
                home_candidates.append(idx)
                
        # Se encontramos candidatos a home
        if home_candidates:
            home_idx = home_candidates[0]  # Usar o primeiro encontrado
            
            # Marcar como Home
            df.at[home_idx, 'Tipo de página'] = "Home"
            df.at[home_idx, 'Visibilidade'] = "Menu"
            df.at[home_idx, 'Hierarquia'] = "Raiz"  # Simplesmente "Raiz" para a home
            
            # Também definir o breadcrumb como "Raiz" se existir essa coluna
            if 'Breadcrumb' in df.columns:
                df.at[home_idx, 'Breadcrumb'] = "Raiz"
            
            logger.info(f"Página Home identificada: {df.at[home_idx, 'De']}")
            logger.info(f"Hierarquia da Home: Raiz")
        else:
            logger.warning("Página Home não identificada.")
            
        return df

    def _process_breadcrumbs(self, df):
        """
        Processa os breadcrumbs para substituir o nome do site por "Raiz".
        """
        logger.info("Processando breadcrumbs...")
        
        # Verificar se temos uma coluna de Breadcrumb
        if 'Breadcrumb' not in df.columns:
            logger.warning("Coluna 'Breadcrumb' não encontrada. Pulando processamento de breadcrumbs.")
            return df
        
        # Fazer uma cópia de segurança dos breadcrumbs originais
        df['Breadcrumb_Original'] = df['Breadcrumb'].copy()
            
        # Substituir o nome do site por "Raiz" em todos os breadcrumbs
        for idx in range(len(df)):
            if pd.notna(df.at[idx, 'Breadcrumb']):
                breadcrumb = df.at[idx, 'Breadcrumb']
                
                # Verificar se começa com o nome do site
                if breadcrumb.startswith(self.site_prefix):
                    # Calcular o novo breadcrumb substituindo o prefixo
                    remainder = breadcrumb[len(self.site_prefix):]
                    new_breadcrumb = "Raiz" + remainder
                    df.at[idx, 'Breadcrumb'] = new_breadcrumb
                    logger.debug(f"Breadcrumb alterado: '{breadcrumb}' -> '{new_breadcrumb}'")
                    
        logger.info("Breadcrumbs processados com sucesso.")
        return df
    
    def _process_hierarchies(self, df):
        """
        Processa as hierarquias, criando uma estrutura lógica.
        Formato desejado: "Raiz > SECAO > SUBSECAO" (substituindo nome do site por "Raiz")
        """
        logger.info("Processando hierarquias...")
        
        # Armazenar hierarquia original antes de modificar
        if 'Hierarquia' in df.columns:
            df['Hierarquia_Original'] = df['Hierarquia'].copy()
        
        # Verificar se temos uma coluna de Breadcrumb para usar
        has_breadcrumb = 'Breadcrumb' in df.columns
        
        # Processar cada linha do DataFrame
        for idx, row in df.iterrows():
            # Pular a Home (já processada)
            if pd.notna(row.get('Tipo de página')) and row.get('Tipo de página') == 'Home':
                continue
                
            # Usar o breadcrumb processado se disponível (já deve estar no formato "Raiz > ...")
            if has_breadcrumb and pd.notna(row['Breadcrumb']):
                df.at[idx, 'Hierarquia'] = row['Breadcrumb']
                continue
            
            # Se não temos o breadcrumb, processar a hierarquia diretamente
            if pd.notna(row['Hierarquia']):
                hierarchy = row['Hierarquia']
                
                # Verificar se começa com o nome do site
                if hierarchy.startswith(self.site_prefix):
                    # Calcular a nova hierarquia substituindo o prefixo
                    remainder = hierarchy[len(self.site_prefix):]
                    new_hierarchy = "Raiz" + remainder
                    df.at[idx, 'Hierarquia'] = new_hierarchy
                    logger.debug(f"Hierarquia alterada: '{hierarchy}' -> '{new_hierarchy}'")
                elif not hierarchy.startswith("Raiz"):
                    # Se não começa com "Raiz" nem com o nome do site, adicionar "Raiz" no início
                    df.at[idx, 'Hierarquia'] = "Raiz > " + hierarchy
                    
        return df
    
    def _process_page_types(self, df):
        """
        Define os tipos de páginas e visibilidade com base em regras específicas.
        """
        logger.info("Processando tipos de páginas e visibilidade...")
        
        # Inicializar campos se não existirem
        if 'Tipo de página' not in df.columns:
            df['Tipo de página'] = None
            
        if 'Visibilidade' not in df.columns:
            df['Visibilidade'] = None
            
        # Processar cada linha
        for idx, row in df.iterrows():
            hierarchy = row['Hierarquia']
            current_type = row['Tipo de página']
            
            # Pular se já for Home (já foi definido anteriormente)
            if pd.notna(current_type) and current_type == 'Home':
                df.at[idx, 'Visibilidade'] = 'Menu'
                continue
                
            # Determinar tipo e visibilidade com base na hierarquia
            parts = hierarchy.split(' > ')
            depth = len(parts)
            
            if depth <= 1:  # Só "Raiz" (caso improvável)
                df.at[idx, 'Tipo de página'] = 'Home'
                df.at[idx, 'Visibilidade'] = 'Menu'
            elif depth == 2:  # Raiz > Seção
                df.at[idx, 'Tipo de página'] = 'Página Definida'
                df.at[idx, 'Visibilidade'] = 'Menu'
            else:  # Níveis mais profundos
                df.at[idx, 'Tipo de página'] = 'Página de Widget'
                df.at[idx, 'Visibilidade'] = 'Oculta'
        
        return df
    
    def _reorder_rows(self, df):
        """
        Reordena as linhas conforme hierarquia para melhor organização,
        agrupando cada página definida com suas respectivas páginas filhas.
        """
        logger.info("Reordenando linhas...")
        
        # Criar uma coluna temporária para a profundidade da hierarquia
        df['depth'] = df['Hierarquia'].apply(lambda x: len(x.split(' > ')) if pd.notna(x) else 999)
        
        # Identificar a Home (deve ficar sempre no topo)
        home_rows = df[df['Tipo de página'] == 'Home'].copy()
        
        # Filtrar todas as linhas que não são a Home
        non_home_rows = df[df['Tipo de página'] != 'Home'].copy()
        
        # Criar dicionário para armazenar grupos de páginas
        hierarchy_groups = {}
        
        # Para cada linha, associamos ao seu parent no nível 2
        for idx, row in non_home_rows.iterrows():
            if pd.notna(row['Hierarquia']):
                hierarchy_parts = row['Hierarquia'].split(' > ')
                
                # Se a hierarquia tiver pelo menos 2 níveis (Raiz > X)
                if len(hierarchy_parts) >= 2:
                    # Usamos o segundo nível como chave do grupo (como TRANSPARÊNCIA)
                    parent_key = hierarchy_parts[1]
                    
                    # Para ordem dentro do grupo, usamos a hierarquia completa
                    # Adicionamos um identificador único para garantir a unicidade
                    if parent_key not in hierarchy_groups:
                        hierarchy_groups[parent_key] = []
                        
                    # A tupla contém: (hierarquia completa, índice original, linha completa)
                    hierarchy_groups[parent_key].append((row['Hierarquia'], idx, row))
                else:
                    # Páginas de nível 1 (apenas Raiz) - improvável, mas possível
                    if 'ROOT' not in hierarchy_groups:
                        hierarchy_groups['ROOT'] = []
                    hierarchy_groups['ROOT'].append((row['Hierarquia'], idx, row))
        
        # Ordenar os grupos de páginas para processamento
        sorted_groups = sorted(hierarchy_groups.keys())
        
        # Criar nova lista para armazenar linhas na ordem correta
        new_rows = []
        
        # Primeiro adicionar a Home
        for _, row in home_rows.iterrows():
            new_rows.append(row)
        
        # Adicionar grupos ordenados (cada página pai seguida de suas filhas)
        for group_key in sorted_groups:
            # Ordenar cada grupo internamente pela hierarquia completa
            items = hierarchy_groups[group_key]
            sorted_items = sorted(items, key=lambda x: (x[0], x[1]))
            
            # Adicionar itens ordenados à nova lista
            for _, _, row in sorted_items:
                new_rows.append(row)
        
        # Criar novo DataFrame com a ordem correta
        new_df = pd.DataFrame(new_rows)
        
        # Remover colunas temporárias
        if 'depth' in new_df.columns:
            new_df = new_df.drop(columns=['depth'])
        
        # Resetar índices
        new_df = new_df.reset_index(drop=True)
        
        return new_df
    
    def _save_to_excel(self, df, filename):
        """
        Salva os dados formatados em um arquivo Excel com estilos.
        """
        logger.info(f"Salvando Excel formatado: {filename}")
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Mapeamento"
            
            # Mesclar células para cabeçalhos principais
            ws.merge_cells('A1:B1')
            ws.merge_cells('C1:F1')
            ws.merge_cells('G1:S1')
            
            # Definição dos grupos e células
            groups = [
                {'title': 'Links', 'range': 'A1:B1', 'color': self.BLUE_COLOR},
                {'title': 'Fase de mapeamento', 'range': 'C1:F1', 'color': self.YELLOW_COLOR},
                {'title': 'Informações sobre a página', 'range': 'G1:S1', 'color': self.RED_COLOR}
            ]
            
            # Aplicar títulos e formatação à linha 1
            for group in groups:
                cell = ws[group['range'].split(':')[0]]
                cell.value = group['title']
                cell.font = Font(bold=True, size=12)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Aplicar cor de fundo a todas as células do grupo
                start_cell, end_cell = group['range'].split(':')
                start_col = ord(start_cell[0]) - ord('A') + 1
                end_col = ord(end_cell[0]) - ord('A') + 1
                
                for col in range(start_col, end_col + 1):
                    ws.cell(row=1, column=col).fill = PatternFill(
                        start_color=group['color'], 
                        end_color=group['color'], 
                        fill_type="solid"
                    )
            
            # Cabeçalhos de segundo nível (usar nomes das colunas do DataFrame)
            headers = list(df.columns)
            # Remover colunas que não queremos exibir
            if 'Hierarquia_Original' in headers:
                headers.remove('Hierarquia_Original')
            if 'Breadcrumb_Original' in headers:
                headers.remove('Breadcrumb_Original')
                
            # Garantir ordem correta das colunas-chave
            key_columns = ['De', 'Para', 'Tipo de migração', 'Qtd de conteúdos', 'Qtd de arquivos', 
                          'Verificar Cópias', 'Hierarquia', 'Visibilidade', 'Menu Lateral']
            
            # Reordenar headers para garantir que as colunas-chave vêm primeiro na ordem correta
            for col in reversed(key_columns):
                if col in headers:
                    headers.remove(col)
                    headers.insert(0, col)
                
            # Aplicar cabeçalhos de segundo nível
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=2, column=col, value=header)
                cell.font = Font(bold=True, size=11)
                cell.alignment = Alignment(horizontal='center', wrap_text=True)
                
                # Aplicar cores de fundo por grupo
                if col <= 2:  # Links (A-B)
                    cell.fill = PatternFill(start_color=self.BLUE_HEADER_COLOR, 
                                           end_color=self.BLUE_HEADER_COLOR, 
                                           fill_type="solid")
                elif col <= 6:  # Fase de mapeamento (C-F)
                    cell.fill = PatternFill(start_color=self.YELLOW_HEADER_COLOR, 
                                           end_color=self.YELLOW_HEADER_COLOR, 
                                           fill_type="solid")
                else:  # Informações sobre a página (G-S)
                    cell.fill = PatternFill(start_color=self.RED_HEADER_COLOR, 
                                           end_color=self.RED_HEADER_COLOR, 
                                           fill_type="solid")
            
            # Thin border style
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Adicionar dados - garantir que usamos apenas as colunas nos headers
            data_df = df[headers].copy()
            
            for r_idx, row in enumerate(data_df.itertuples(), 3):
                for c_idx, value in enumerate(row[1:], 1):  # Skip index
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    cell.border = thin_border
                    
                    # Aplicar cores de fundo alternadas por grupo
                    if c_idx <= 2:  # Links
                        cell.fill = PatternFill(start_color=self.BLUE_LIGHT_COLOR, 
                                               end_color=self.BLUE_LIGHT_COLOR, 
                                               fill_type="solid")
                    elif c_idx <= 6:  # Fase de mapeamento
                        cell.fill = PatternFill(start_color=self.YELLOW_LIGHT_COLOR, 
                                               end_color=self.YELLOW_LIGHT_COLOR, 
                                               fill_type="solid")
                    else:  # Informações sobre a página
                        cell.fill = PatternFill(start_color=self.RED_LIGHT_COLOR, 
                                               end_color=self.RED_LIGHT_COLOR, 
                                               fill_type="solid")
            
            # Adicionar fórmulas para verificar duplicados na coluna F (Verificar Cópias)
            for row in range(3, ws.max_row + 1):
                cell = ws.cell(row=row, column=6)
                cell.value = f'=IF(COUNTIF(A:A,A{row})>1,"Duplicado","Único")'
            
            # Ajustar largura das colunas
            for col_idx in range(1, ws.max_column + 1):
                max_length = 0
                column_letter = get_column_letter(col_idx)
                
                for row_idx in range(1, ws.max_row + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if hasattr(cell, 'value') and cell.value:
                        try:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                        except:
                            pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Congelar painel nas duas primeiras linhas
            ws.freeze_panes = 'A3'
            
            # Configurar altura da linha 1 um pouco maior para melhor visualização
            ws.row_dimensions[1].height = 24
            
            # Salvar o arquivo
            wb.save(filename)
            logger.info(f"Excel salvo com sucesso: {filename}")
            
        except Exception as e:
            logger.error(f"Erro ao salvar Excel: {str(e)}", exc_info=True)
            raise
