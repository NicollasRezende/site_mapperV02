from datetime import datetime
from typing import Dict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import logging
from models.page_data import PageData

logger = logging.getLogger(__name__)

class ExcelService:
    @staticmethod
    def save_to_excel(pages: Dict[str, PageData], test_mode: bool = False):
        """
        Salva dados das páginas em um arquivo Excel com layout melhorado.
        
        Args:
            pages: Dicionário das páginas mapeadas
            test_mode: Se True, processa apenas as primeiras 30 páginas para teste
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"logs/mapeamento_site_{timestamp}.xlsx"
        
        logger.info(f"Salvando dados em Excel: {filename}")
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Mapeamento"
            
            # Cores definidas - ATUALIZADAS conforme solicitado
            BLUE_COLOR = "9FC5E8"
            YELLOW_COLOR = "FFE599"
            RED_COLOR = "EA9999"
            
            # Cores para linha 2 - ATUALIZADAS conforme solicitado
            BLUE_HEADER_COLOR = "CFE2F3"    # Linha 2A-2B
            YELLOW_HEADER_COLOR = "FFFAEC"  # Linha 2C-2F
            RED_HEADER_COLOR = "F4CCCC"     # Linha 2G-2S
            
            # Cores para dados - cores claras para melhor leitura
            BLUE_LIGHT_COLOR = "EDF5FC"
            YELLOW_LIGHT_COLOR = "FFFAEC"
            RED_LIGHT_COLOR = "FFF2F2"
            
            # Estilos
            header_font = Font(bold=True, color="000000")
            
            # Borda
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Definir as colunas e grupos
            # Linha 1 - Mesclar células e adicionar cabeçalhos principais
            ws.merge_cells('A1:B1')
            ws.merge_cells('C1:F1')
            ws.merge_cells('G1:S1')
            
            # Definição dos grupos e células
            groups = [
                {'title': 'Links', 'range': 'A1:B1', 'color': BLUE_COLOR},
                {'title': 'Fase de mapeamento', 'range': 'C1:F1', 'color': YELLOW_COLOR},
                {'title': 'Informações sobre a página', 'range': 'G1:S1', 'color': RED_COLOR}
            ]
            
            # Aplicar títulos e formatação à linha 1
            for group in groups:
                cell = ws[group['range'].split(':')[0]]
                cell.value = group['title']
                # Garantir que o texto da linha 1 esteja em negrito
                cell.font = Font(bold=True, size=12)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Aplicar cor de fundo a todas as células do grupo
                start_cell, end_cell = group['range'].split(':')
                start_col = ord(start_cell[0]) - ord('A') + 1
                end_col = ord(end_cell[0]) - ord('A') + 1
                
                for col in range(start_col, end_col + 1):
                    ws.cell(row=1, column=col).fill = PatternFill(start_color=group['color'], 
                                                              end_color=group['color'], 
                                                              fill_type="solid")
            
            # Linha 2 - Cabeçalhos de segundo nível
            headers_row2 = [
                # Links
                'De', 'Para',
                # Fase de mapeamento
                'Tipo de migração', 'Qtd de conteúdos', 'Qtd de arquivos', 'Verificar Cópias',
                # Informações sobre a página
                'Hierarquia', 'Visibilidade', 'Menu Lateral', 'Breadcrumb', 'Vocabulário',
                'Categoria', 'Pontos de atenção', 'Redes sociais', 'Tipo de página',
                '(Para página de vincular a uma página deste site) Nome da página',
                'Link da página para a qual redireciona', 'Complexidade', 'Layout'
            ]
            
            # Aplicar cabeçalhos de segundo nível
            for col, header in enumerate(headers_row2, 1):
                cell = ws.cell(row=2, column=col, value=header)
                # Garantir que o texto da linha 2 esteja em negrito
                cell.font = Font(bold=True, size=11)
                cell.alignment = Alignment(horizontal='center', wrap_text=True)
                
                # Aplicar cores de fundo por grupo - CORES ATUALIZADAS
                if col <= 2:  # Links (A-B)
                    cell.fill = PatternFill(start_color=BLUE_HEADER_COLOR, end_color=BLUE_HEADER_COLOR, fill_type="solid")
                elif col <= 6:  # Fase de mapeamento (C-F)
                    cell.fill = PatternFill(start_color=YELLOW_HEADER_COLOR, end_color=YELLOW_HEADER_COLOR, fill_type="solid")
                else:  # Informações sobre a página (G-S)
                    cell.fill = PatternFill(start_color=RED_HEADER_COLOR, end_color=RED_HEADER_COLOR, fill_type="solid")
            
            # Processar e inserir os dados das páginas
            row = 3
            
            # Limitar páginas se estiver em modo de teste
            page_items = list(pages.items())
            if test_mode:
                # CORRIGIDO: No modo de teste, limitamos a 30 páginas
                page_items = page_items[:30]
                logger.info("Modo de teste ativado - processando apenas 30 páginas")
            else:
                # CORRIGIDO: Agora processamos todas as páginas quando não estiver em modo de teste
                logger.info(f"Modo completo ativado - processando todas as {len(page_items)} páginas")
            
            for url, page in page_items:
                row_data = page.to_planilha_row()
                
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = thin_border
                    
                    # Aplicar cores de fundo alternadas por grupo
                    if col <= 2:  # Links
                        cell.fill = PatternFill(start_color=BLUE_LIGHT_COLOR, end_color=BLUE_LIGHT_COLOR, fill_type="solid")
                    elif col <= 6:  # Fase de mapeamento
                        cell.fill = PatternFill(start_color=YELLOW_LIGHT_COLOR, end_color=YELLOW_LIGHT_COLOR, fill_type="solid")
                    else:  # Informações sobre a página
                        cell.fill = PatternFill(start_color=RED_LIGHT_COLOR, end_color=RED_LIGHT_COLOR, fill_type="solid")
                
                row += 1

            # Ajustar largura das colunas - CORREÇÃO: tratar células mescladas
            for col_idx in range(1, ws.max_column + 1):
                max_length = 0
                column_letter = get_column_letter(col_idx)
                
                for row_idx in range(1, ws.max_row + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if hasattr(cell, 'value') and cell.value:
                        try:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                        except:
                            pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Congelar painel nas duas primeiras linhas
            ws.freeze_panes = 'A3'
            
            # Configurar altura da linha 1 um pouco maior para melhor visualização
            ws.row_dimensions[1].height = 24
            
            wb.save(filename)
            logger.info(f"Arquivo Excel salvo com sucesso: {filename}")
            logger.info(f"Total de páginas salvas: {len(page_items)}")
            
        except Exception as e:
            logger.error(f"Erro ao salvar arquivo Excel: {e}", exc_info=True)
            raise